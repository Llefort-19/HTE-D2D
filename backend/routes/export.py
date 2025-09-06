"""
Export routes blueprint.
Handles experiment data export to Excel format.
"""
import os
import tempfile
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from state import current_experiment, inventory_data

# Create blueprint
export_bp = Blueprint('export', __name__, url_prefix='/api/experiment')

@export_bp.route('/export', methods=['POST'])
def export_experiment():
    """Export experiment data to Excel format"""
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Context sheet
    ws_context = wb.create_sheet("Context")
    context_data = [
        ['Author', current_experiment['context'].get('author', '')],
        ['Date', current_experiment['context'].get('date', '')],
        ['Project', current_experiment['context'].get('project', '')],
        ['ELN', current_experiment['context'].get('eln', '')],
        ['Objective', current_experiment['context'].get('objective', '')]
    ]
    
    for row in context_data:
        ws_context.append(row)
    
    # Materials sheet
    ws_materials = wb.create_sheet("Materials")
    if current_experiment.get('materials'):
        # Add headers - match inventory column names exactly (lowercase) and order
        headers = ['Nr', 'chemical_name', 'alias', 'cas_number', 'molecular_weight', 'smiles', 'barcode', 'role', 'source', 'supplier']
        ws_materials.append(headers)
        
        # Load inventory data to enrich materials
        inventory_enrichment = {}
        if inventory_data:
            for _, inv_item in inventory_data.iterrows():
                # Create lookup keys for matching
                name_key = str(inv_item.get('chemical_name', '')).lower()
                cas_key = str(inv_item.get('cas_number', '')).lower()
                alias_key = str(inv_item.get('alias', '')).lower()
                
                # Store inventory data for matching
                inventory_enrichment[name_key] = inv_item.to_dict()
                if cas_key and cas_key != 'nan':
                    inventory_enrichment[cas_key] = inv_item.to_dict()
                if alias_key and alias_key != 'nan':
                    inventory_enrichment[alias_key] = inv_item.to_dict()
        
        # Add materials with enriched data from inventory
        for i, material in enumerate(current_experiment['materials'], 1):
            # Try to find matching inventory data
            enriched_data = {}
            material_name = str(material.get('name', '')).lower()
            material_cas = str(material.get('cas', '')).lower()
            material_alias = str(material.get('alias', '')).lower()
            
            # Look for matches in inventory
            if material_name in inventory_enrichment:
                enriched_data = inventory_enrichment[material_name]
            elif material_cas in inventory_enrichment and material_cas != 'nan':
                enriched_data = inventory_enrichment[material_cas]
            elif material_alias in inventory_enrichment and material_alias != 'nan':
                enriched_data = inventory_enrichment[material_alias]
            
            # Use material data first, then enrich with inventory data
            row = [
                i,
                material.get('name', ''),
                material.get('alias', enriched_data.get('alias', '')),
                material.get('cas', enriched_data.get('cas_number', '')),
                material.get('molecular_weight', enriched_data.get('molecular_weight', '')),
                material.get('smiles', enriched_data.get('smiles', '')),
                material.get('barcode', enriched_data.get('barcode', '')),
                material.get('role', ''),
                material.get('source', enriched_data.get('source', '')),
                material.get('supplier', enriched_data.get('supplier', ''))
            ]
            ws_materials.append(row)
    
    # Procedure sheet
    ws_procedure = wb.create_sheet("Procedure")
    if current_experiment.get('procedure'):
        # Add headers for 96-well plate
        headers = ['Nr', 'Well', 'ID']
        # Add compound columns (up to 15 compounds)
        for i in range(1, 16):
            headers.extend([f'Compound-{i}_name', f'Compound-{i}_mmol'])
        # Add reagent columns (up to 5 reagents)
        for i in range(1, 6):
            headers.extend([f'Reagent-{i}_name', f'Reagent-{i}_mmol'])
        # Add solvent columns (up to 3 solvents)
        for i in range(1, 4):
            headers.extend([f'Solvent-{i}_name', f'Solvent-{i}_uL'])
        
        ws_procedure.append(headers)
        
        # Add procedure data
        for i, well_data in enumerate(current_experiment['procedure'], 1):
            row = [i, well_data.get('well', ''), well_data.get('id', '')]
            
            # Add compounds
            for j in range(1, 16):
                row.extend([
                    well_data.get(f'compound_{j}_name', ''),
                    well_data.get(f'compound_{j}_mmol', '')
                ])
            
            # Add reagents
            for j in range(1, 6):
                row.extend([
                    well_data.get(f'reagent_{j}_name', ''),
                    well_data.get(f'reagent_{j}_mmol', '')
                ])
            
            # Add solvents
            for j in range(1, 4):
                row.extend([
                    well_data.get(f'solvent_{j}_name', ''),
                    well_data.get(f'solvent_{j}_ul', '')
                ])
            
            ws_procedure.append(row)
    
    # Analytical data sheet
    ws_analytical = wb.create_sheet("Analytical data (1)")
    analytical_data = current_experiment.get('analytical_data', {})
    
    # Handle both old format (list) and new format (dict with uploadedFiles)
    if analytical_data:
        # Check if it's the new format
        if isinstance(analytical_data, dict) and 'uploadedFiles' in analytical_data:
            uploaded_files = analytical_data.get('uploadedFiles', [])
            if uploaded_files:
                # Use the first uploaded file's data
                file_data = uploaded_files[0].get('data', [])
                if file_data:
                    # Add headers based on the actual data structure
                    headers = list(file_data[0].keys()) if file_data else ['Well', 'Sample ID']
                    ws_analytical.append(headers)
                    
                    # Add data rows
                    for row_data in file_data:
                        row = [row_data.get(key, '') for key in headers]
                        ws_analytical.append(row)
        elif isinstance(analytical_data, list):
            # Handle old format (list of analytical data)
            if analytical_data:
                # Add headers
                headers = ['Nr', 'Well', 'ID']
                for i in range(1, 16):
                    headers.extend([f'Compound-{i}_name', f'Compound-{i}_area'])
                
                ws_analytical.append(headers)
                
                # Add analytical data
                for i, data_item in enumerate(analytical_data, 1):
                    if isinstance(data_item, dict):
                        row = [i, data_item.get('well', ''), data_item.get('id', '')]
                        
                        for j in range(1, 16):
                            row.extend([
                                data_item.get(f'compound_{j}_name', ''),
                                data_item.get(f'compound_{j}_area', '')
                            ])
                        
                        ws_analytical.append(row)
    
    # Results sheet
    ws_results = wb.create_sheet("Results (1)")
    if current_experiment.get('results'):
        # Add headers
        headers = ['Nr', 'Well', 'ID', 'Conversion_%', 'Yield_%', 'Selectivity_%']
        ws_results.append(headers)
        
        # Add results data
        for i, result_data in enumerate(current_experiment['results'], 1):
            row = [
                i,
                result_data.get('well', ''),
                result_data.get('id', ''),
                result_data.get('conversion_percent', ''),
                result_data.get('yield_percent', ''),
                result_data.get('selectivity_percent', '')
            ]
            ws_results.append(row)
    
    # Well Contents sheet - Detailed view of each well
    ws_well_contents = wb.create_sheet("Well Contents")
    
    # Create a mapping of materials by name for quick lookup
    materials_map = {}
    for material in current_experiment.get('materials', []):
        materials_map[material.get('name', '').lower()] = material
    
    # Initialize well contents data
    well_contents = {}
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for row in range(1, 13):
            well = f'{col}{row}'
            well_contents[well] = {
                'compounds': [],
                'reagents': [],
                'solvents': []
            }
    
    # Fill in well contents from procedure data
    if current_experiment.get('procedure'):
        for well_data in current_experiment['procedure']:
            well = well_data.get('well', '')
            if well in well_contents:
                # Process materials array
                materials = well_data.get('materials', [])
                
                for material in materials:
                    name = material.get('name', '')
                    amount = material.get('amount', '')
                    alias = material.get('alias', '')
                    cas = material.get('cas', '')
                    
                    if name and amount:
                        # For now, treat all materials as compounds
                        # You can add logic here to distinguish compounds, reagents, solvents
                        well_contents[well]['compounds'].append({
                            'name': name,
                            'amount': amount,
                            'alias': alias,
                            'cas': cas
                        })
    
    # Find the maximum number of compounds across all wells to determine column count
    max_compounds = 0
    for well in well_contents:
        compounds_count = len(well_contents[well]['compounds'])
        reagents_count = len(well_contents[well]['reagents'])
        solvents_count = len(well_contents[well]['solvents'])
        total_materials = compounds_count + reagents_count + solvents_count
        max_compounds = max(max_compounds, total_materials)
    
    # Create header row
    headers = ['Well']
    for i in range(1, max_compounds + 1):
        headers.extend([f'Compound_{i}_Name', f'Compound_{i}_Alias', f'Compound_{i}_CAS', f'Compound_{i}_Amount'])
    
    ws_well_contents.append(headers)
    
    # Add data for each well (all 96 wells)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for row in range(1, 13):
            well = f'{col}{row}'
            contents = well_contents[well]
            
            # Combine all materials into a single list
            all_materials = []
            all_materials.extend(contents['compounds'])
            all_materials.extend(contents['reagents'])
            all_materials.extend(contents['solvents'])
            
            # Create row data
            row_data = [well]
            
            # Add materials to columns (4 columns per material)
            for i in range(max_compounds):
                if i < len(all_materials):
                    material = all_materials[i]
                    row_data.extend([
                        material.get('name', ''),
                        material.get('alias', ''),
                        material.get('cas', ''),
                        material.get('amount', '')
                    ])
                else:
                    # Fill empty columns
                    row_data.extend(['', '', '', ''])
            
            ws_well_contents.append(row_data)
    
    # Procedure Settings sheet
    ws_procedure_settings = wb.create_sheet("Procedure Settings")
    
    # Reaction Conditions section
    ws_procedure_settings.append(['Reaction Conditions'])
    ws_procedure_settings.append(['Parameter', 'Value', 'Unit'])
    ws_procedure_settings.append(['Temperature', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('temperature', ''), 'degC'])
    ws_procedure_settings.append(['Time', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('time', ''), 'h'])
    ws_procedure_settings.append(['Pressure', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('pressure', ''), 'bar'])
    ws_procedure_settings.append(['Wavelength', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('wavelength', ''), 'nm'])
    ws_procedure_settings.append([''])  # Empty row for spacing
    ws_procedure_settings.append(['Remarks'])
    ws_procedure_settings.append([current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('remarks', '')])
    
    # Analytical Details section
    ws_procedure_settings.append([''])  # Empty row for spacing
    ws_procedure_settings.append(['Analytical Details'])
    ws_procedure_settings.append(['Parameter', 'Value', 'Unit'])
    ws_procedure_settings.append(['UPLC #', current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('uplcNumber', ''), ''])
    ws_procedure_settings.append(['Method', current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('method', ''), ''])
    ws_procedure_settings.append(['Duration', current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('duration', ''), 'min'])
    ws_procedure_settings.append([''])  # Empty row for spacing
    ws_procedure_settings.append(['Remarks'])
    ws_procedure_settings.append([current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('remarks', '')])
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    # Generate filename based on ELN number or timestamp
    context = current_experiment.get('context', {})
    eln_number = context.get('eln', '').strip()
    
    if eln_number:
        # Use ELN number + date (YYYY-MM-DD format)
        date_only = datetime.now().strftime("%Y-%m-%d")
        filename = f'{eln_number}_{date_only}.xlsx'
    else:
        # Fallback to original timestamp format
        filename = f'HTE_experiment_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(tmp_path, as_attachment=True, download_name=filename)
