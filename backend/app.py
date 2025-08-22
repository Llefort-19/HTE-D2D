# Standard library imports
import os
import io
import base64
import tempfile
from datetime import datetime
import re

# Third-party imports
import pandas as pd
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image

# Chemical informatics imports
try:
    from rdkit import Chem
    from rdkit.Chem import Draw, AllChem
    RDKIT_AVAILABLE = True
except ImportError:
    print("Warning: RDKit not available. Molecule rendering will be disabled.")
    RDKIT_AVAILABLE = False

app = Flask(__name__)
CORS(app)

# Global variables to store data
inventory_data = None
current_experiment = {
    'context': {},
    'materials': [],
    'procedure': [],
    'procedure_settings': {
        'reactionConditions': {
            'temperature': '',
            'time': '',
            'pressure': '',
            'wavelength': '',
            'remarks': ''
        },
        'analyticalDetails': {
            'uplcNumber': '',
            'method': '',
            'duration': '',
            'remarks': ''
        }
    },
    'analytical_data': {
        'selectedCompounds': [],
        'uploadedFiles': []
    },
    'results': []
}

def image_to_base64(img_or_bytes):
    """Convert PIL.Image or PNG bytes to base64 string."""
    try:
        if hasattr(img_or_bytes, 'save'):
            buffer = io.BytesIO()
            img_or_bytes.save(buffer, format='PNG')
            img_bytes = buffer.getvalue()
        else:
            img_bytes = img_or_bytes
        return base64.b64encode(img_bytes).decode()
    except Exception as e:
        print(f"[image_to_base64] Error: {e}")
        return None

def blank_png_base64(size=(300, 300)):
    """Generate a blank PNG image as base64 string."""
    img = Image.new("RGBA", size, (255, 255, 255, 0))
    return image_to_base64(img)

def normalize_2d_coordinates(mol):
    """Normalize 2D coordinates for consistent rendering."""
    if not RDKIT_AVAILABLE:
        return mol
    
    try:
        AllChem.Compute2DCoords(mol)
        return mol
    except Exception as e:
        print(f"[normalize_2d_coordinates] Error: {e}")
        return mol

def prepare_molecule(smiles_string):
    """Prepare molecule from SMILES string."""
    if not RDKIT_AVAILABLE:
        return None
    
    try:
        mol = Chem.MolFromSmiles(smiles_string.strip())
        if mol is None:
            print(f"[prepare_molecule] Invalid SMILES: {smiles_string}")
            return None
        
        mol = normalize_2d_coordinates(mol)
        return mol
    except Exception as e:
        print(f"[prepare_molecule] Error: {e}")
        return None

def render_molecule_png(mol, image_size=(300, 300)):
    """Render molecule as PNG bytes using RDKit's built-in PIL renderer."""
    if not RDKIT_AVAILABLE or mol is None:
        return None
    
    try:
        # Use RDKit's PIL-based PNG drawer (no cairosvg needed)
        img = Draw.MolToImage(mol, size=image_size)
        
        # Convert PIL image to PNG bytes
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        png_bytes = img_buffer.getvalue()
        
        return png_bytes
    except Exception as e:
        print(f"[render_molecule_png] Error: {e}")
        return None

def generate_molecule_image(smiles_string, image_size=(300, 300)):
    """
    Generate a 2D molecule image from a SMILES string.
    Returns: base64 encoded PNG image or None if error.
    """
    if not RDKIT_AVAILABLE:
        print("[generate_molecule_image] RDKit not available")
        return blank_png_base64(image_size)
    
    try:
        mol = prepare_molecule(smiles_string)
        if mol is None:
            print(f"[generate_molecule_image] Could not prepare molecule from: {smiles_string}")
            return blank_png_base64(image_size)
        
        png_bytes = render_molecule_png(mol, image_size)
        if png_bytes:
            return image_to_base64(png_bytes)
        else:
            print(f"[generate_molecule_image] Could not render PNG for: {smiles_string}")
            return blank_png_base64(image_size)
    except Exception as e:
        print(f"[generate_molecule_image] Error with {smiles_string}: {e}")
        return blank_png_base64(image_size)

def parse_sdf_file(sdf_content):
    """
    Parse SDF file content and extract molecules with images.
    
    Args:
        sdf_content (str): The content of the SDF file
    
    Returns:
        list: List of molecule dictionaries with name, smiles, and image
    """
    if not RDKIT_AVAILABLE:
        print("[parse_sdf_file] RDKit not available")
        return []
    
    molecules = []
    
    try:
        print(f"[parse_sdf_file] Processing SDF content, length: {len(sdf_content)}")
        
        # Use RDKit to parse SDF
        mol_supplier = Chem.SDMolSupplier()
        mol_supplier.SetData(sdf_content)
        
        for i, mol in enumerate(mol_supplier):
            if mol is None:
                print(f"[parse_sdf_file] Skipping invalid molecule at index {i}")
                continue
            
            try:
                # Get molecule name from SDF properties or generate one
                mol_name = mol.GetProp('_Name') if mol.HasProp('_Name') else f"Molecule_{i+1}"
                
                # Generate SMILES
                smiles = Chem.MolToSmiles(mol)
                
                # Generate molecule image
                image_size = (200, 200)  # Smaller size for table display
                mol_2d = normalize_2d_coordinates(mol)
                png_bytes = render_molecule_png(mol_2d, image_size)
                
                image_base64 = None
                if png_bytes:
                    image_base64 = image_to_base64(png_bytes)
                
                molecule_data = {
                    'name': mol_name,
                    'smiles': smiles,
                    'image': image_base64,
                    'role': ''  # Will be set by user
                }
                
                molecules.append(molecule_data)
                print(f"[parse_sdf_file] Processed molecule {i+1}: {mol_name}")
                
            except Exception as e:
                print(f"[parse_sdf_file] Error processing molecule {i+1}: {e}")
                continue
        
        print(f"[parse_sdf_file] Successfully processed {len(molecules)} molecules")
        return molecules
        
    except Exception as e:
        print(f"[parse_sdf_file] Error parsing SDF: {e}")
        return []

def load_inventory():
    """Load inventory from Excel file"""
    global inventory_data
    try:
        inventory_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Inventory.xlsx')
        # Read Excel file without parsing dates to avoid NaTType issues
        inventory_data = pd.read_excel(inventory_path, parse_dates=False)
        
        # Convert all columns to string to avoid any datetime/NaT issues
        for col in inventory_data.columns:
            inventory_data[col] = inventory_data[col].astype(str)
            # Replace 'nan' strings with None for better JSON handling
            inventory_data[col] = inventory_data[col].replace('nan', None)
        
        return True
    except Exception as e:
        print(f"Error loading inventory: {e}")
        return False

@app.route('/api/inventory', methods=['GET'])
def get_inventory():
    """Get all chemicals from inventory"""
    if inventory_data is None:
        if not load_inventory():
            return jsonify({'error': 'Failed to load inventory'}), 500
    
    if inventory_data is not None:
        # Clean the data before JSON serialization to handle NaT values
        records = inventory_data.to_dict('records')
        cleaned_records = []
        
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                # Handle pandas NaT values and other problematic types
                if pd.isna(value) or (hasattr(value, 'year') and pd.isna(value)):
                    cleaned_record[key] = None
                elif hasattr(value, 'isoformat'):  # datetime objects
                    cleaned_record[key] = value.isoformat()
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)
        
        return jsonify(cleaned_records)
    else:
        return jsonify([])

@app.route('/api/inventory/search', methods=['GET'])
def search_inventory():
    """Search chemicals in both main and private inventory"""
    query = request.args.get('q', '').lower()
    
    # Main inventory
    if inventory_data is None:
        if not load_inventory():
            return jsonify({'error': 'Failed to load inventory'}), 500
    
    main_results = pd.DataFrame()
    if inventory_data is not None:
        main_results = inventory_data[
            inventory_data['chemical_name'].str.lower().str.contains(query, na=False) |
            inventory_data['alias'].str.lower().str.contains(query, na=False) |
            inventory_data['cas_number'].astype(str).str.lower().str.contains(query, na=False) |
            inventory_data['smiles'].astype(str).str.lower().str.contains(query, na=False)
        ]
    
    # Private inventory
    private_path = os.path.join(os.path.dirname(__file__), '..', 'Private_Inventory.xlsx')
    private_results = pd.DataFrame()
    if os.path.exists(private_path):
        try:
            # Read private inventory without parsing dates to avoid NaTType issues
            private_df = pd.read_excel(private_path, parse_dates=False)
            
            # Convert all columns to string to avoid any datetime/NaT issues
            for col in private_df.columns:
                private_df[col] = private_df[col].astype(str)
                # Replace 'nan' strings with None for better JSON handling
                private_df[col] = private_df[col].replace('nan', None)
            
            private_results = private_df[
                private_df['chemical_name'].str.lower().str.contains(query, na=False) |
                private_df['alias'].str.lower().str.contains(query, na=False) |
                private_df['cas_number'].astype(str).str.lower().str.contains(query, na=False) |
                private_df['smiles'].astype(str).str.lower().str.contains(query, na=False)
            ]
        except Exception as e:
            print(f"Error loading private inventory: {e}")
            pass
    
    # Combine with main inventory priority
    if not main_results.empty and not private_results.empty:
        # Get names and CAS from main results to filter out duplicates from private
        main_names = set(main_results['chemical_name'].str.lower())
        main_cas = set(main_results['cas_number'].astype(str).str.lower())
        
        # Filter private results to exclude duplicates
        private_filtered = private_results[
            ~(private_results['chemical_name'].str.lower().isin(main_names) |
              private_results['cas_number'].astype(str).str.lower().isin(main_cas))
        ]
        
        # Combine main results with filtered private results
        combined = pd.concat([main_results, private_filtered], ignore_index=True)
    elif not main_results.empty:
        combined = main_results
    elif not private_results.empty:
        combined = private_results
    else:
        combined = pd.DataFrame()
    
    # Clean the data before JSON serialization to handle NaT values
    if not combined.empty:
        # Convert DataFrame to records and clean any problematic values
        records = combined.to_dict('records')
        cleaned_records = []
        
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                # Handle pandas NaT values and other problematic types
                if pd.isna(value) or (hasattr(value, 'year') and pd.isna(value)):
                    cleaned_record[key] = None
                elif hasattr(value, 'isoformat'):  # datetime objects
                    cleaned_record[key] = value.isoformat()
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)
        
        return jsonify(cleaned_records)
    else:
        return jsonify([])

@app.route('/api/solvent/search', methods=['GET'])
def search_solvents():
    """Search solvents in the Solvent.xlsx file"""
    query = request.args.get('q', '').lower()
    search_type = request.args.get('type', 'all')  # all, name, alias, cas, boiling_point, class
    class_filter = request.args.get('class_filter', '').lower()
    bp_filter = request.args.get('bp_filter', '')
    tier_filter = request.args.get('tier_filter', '')
    
    solvent_path = os.path.join(os.path.dirname(__file__), '..', 'Solvent.xlsx')
    
    if not os.path.exists(solvent_path):
        return jsonify({'error': 'Solvent database not found'}), 404
    
    try:
        df = pd.read_excel(solvent_path)
        
        # Handle NaN values
        df = df.fillna('')
        
        # Start with all data
        results = df.copy()
        
        # Apply text search if query provided
        if query:
            text_filter = (
                df['Name'].astype(str).str.lower().str.contains(query, na=False) |
                df['Alias'].astype(str).str.lower().str.contains(query, na=False) |
                df['CAS Number'].astype(str).str.lower().str.contains(query, na=False)
            )
            results = results[text_filter]
            print(f"Text filter results: {len(results)} matches found")
        
        # Apply class filter if provided
        if class_filter:
            print(f"Applying class filter: '{class_filter}'")
            # More flexible class matching - check if the class filter is contained in the chemical class
            # Also handle common variations and plural forms
            class_variations = [class_filter]
            if class_filter.endswith('s'):
                class_variations.append(class_filter[:-1])  # Remove 's' for singular
            else:
                class_variations.append(class_filter + 's')  # Add 's' for plural
            
            # Create a more flexible filter
            class_mask = results['Chemical Class'].astype(str).str.lower().str.contains('|'.join(class_variations), na=False)
            print(f"Class filter results: {class_mask.sum()} matches found")
            print(f"Available classes in filtered data: {results['Chemical Class'].astype(str).unique()}")
            results = results[class_mask]
        
        # Apply boiling point filter if provided
        if bp_filter:
            try:
                if bp_filter.startswith('>'):
                    bp_value = float(bp_filter[1:].strip())
                    bp_mask = results['Boiling point'] > bp_value
                elif bp_filter.startswith('<'):
                    bp_value = float(bp_filter[1:].strip())
                    bp_mask = results['Boiling point'] < bp_value
                else:
                    # Try to parse as exact value
                    bp_value = float(bp_filter)
                    tolerance = 5  # ±5°C tolerance
                    bp_mask = (results['Boiling point'] >= bp_value - tolerance) & (results['Boiling point'] <= bp_value + tolerance)
                
                print(f"Boiling point filter results: {bp_mask.sum()} matches found")
                results = results[bp_mask]
            except ValueError:
                # If boiling point filter is invalid, return empty results
                print("Invalid boiling point filter value")
                results = pd.DataFrame()
        
        # Apply tier filter if provided
        if tier_filter:
            try:
                max_tier = int(tier_filter)
                # Extract numeric part from "Tier X" format
                tier_numeric = results['Tier'].astype(str).str.extract(r'Tier\s*(\d+)')[0].astype(float)
                tier_mask = tier_numeric <= max_tier
                results = results[tier_mask]
            except ValueError:
                # If tier filter is invalid, return empty results
                results = pd.DataFrame()
        
        # Convert to list of dictionaries with consistent field names
        solvent_results = []
        for _, row in results.iterrows():
            solvent_results.append({
                'name': row['Name'],
                'alias': row['Alias'],
                'cas': row['CAS Number'],
                'molecular_weight': row['Molecular_weight'],
                'smiles': row['SMILES'],
                'boiling_point': row['Boiling point'],
                'chemical_class': row['Chemical Class'],
                'density': row['Density (g/mL)'],
                'tier': row['Tier'],
                'source': 'solvent_database'
            })
        
        return jsonify(solvent_results)
        
    except Exception as e:
        return jsonify({'error': f'Error searching solvents: {str(e)}'}), 500

@app.route('/api/solvent/tiers', methods=['GET'])
def get_solvent_tiers():
    """Get all available solvent tiers from the database"""
    solvent_path = os.path.join(os.path.dirname(__file__), '..', 'Solvent.xlsx')
    
    if not os.path.exists(solvent_path):
        return jsonify({'error': 'Solvent database not found'}), 404
    
    try:
        df = pd.read_excel(solvent_path)
        df = df.fillna('')
        
        # Get unique tiers
        tiers = df['Tier'].astype(str).unique()
        tiers = [tier.strip() for tier in tiers if tier.strip() and tier.strip().lower() != 'nan']
        
        # Extract numeric part from "Tier X" format and convert to integers
        tier_numbers = []
        for tier in tiers:
            try:
                # Extract number from "Tier X" format
                match = re.search(r'Tier\s*(\d+)', tier, re.IGNORECASE)
                if match:
                    tier_numbers.append(int(match.group(1)))
            except (ValueError, AttributeError):
                continue
        
        # Sort and convert back to strings
        tier_numbers.sort()
        tiers = [str(tier) for tier in tier_numbers]
        
        return jsonify(tiers)
        
    except Exception as e:
        return jsonify({'error': f'Error getting solvent tiers: {str(e)}'}), 500

@app.route('/api/solvent/classes', methods=['GET'])
def get_solvent_classes():
    """Get all available solvent classes from the database"""
    solvent_path = os.path.join(os.path.dirname(__file__), '..', 'Solvent.xlsx')
    
    if not os.path.exists(solvent_path):
        return jsonify({'error': 'Solvent database not found'}), 404
    
    try:
        df = pd.read_excel(solvent_path)
        df = df.fillna('')
        
        # Get unique chemical classes
        classes = df['Chemical Class'].astype(str).unique()
        classes = [cls.strip() for cls in classes if cls.strip() and cls.strip().lower() != 'nan']
        
        # Sort classes alphabetically
        classes.sort()
        
        return jsonify(classes)
        
    except Exception as e:
        return jsonify({'error': f'Error getting solvent classes: {str(e)}'}), 500

@app.route('/api/inventory/private/add', methods=['POST'])
def add_to_private_inventory():
    chemical = request.json
    private_path = os.path.join(os.path.dirname(__file__), '..', 'Private_Inventory.xlsx')
    headers = ['chemical_name', 'alias', 'cas_number', 'molecular_weight', 'smiles', 'barcode']

    # Create file if it doesn't exist
    if not os.path.exists(private_path):
        wb = Workbook()
        # Remove the default sheet and create a new one with the correct name
        wb.remove(wb.active)
        ws = wb.create_sheet("Private Inventory")
        ws.append(headers)
        wb.save(private_path)

    # Load and check for duplicates
    df = pd.read_excel(private_path)
    
    # Ensure the DataFrame has only the correct columns
    required_columns = ['chemical_name', 'alias', 'cas_number', 'molecular_weight', 'smiles', 'barcode']
    for col in required_columns:
        if col not in df.columns:
            df[col] = ''
    
    # Remove any extra columns that shouldn't be there
    df = df[required_columns]
    
    if ((df['chemical_name'].str.lower() == chemical['name'].lower()) | 
        (df['cas_number'].astype(str) == str(chemical.get('cas', '')))).any():
        return jsonify({'message': 'Already exists'}), 200

    # Append and save
    new_row = {
        'chemical_name': chemical['name'],
        'alias': chemical.get('alias', ''),
        'cas_number': chemical.get('cas', ''),
        'molecular_weight': chemical.get('molecular_weight', ''),
        'smiles': chemical.get('smiles', ''),
        'barcode': chemical.get('barcode', '')
    }
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_excel(private_path, index=False)
    return jsonify({'message': 'Added'}), 200

@app.route('/api/inventory/private/fix-structure', methods=['POST'])
def fix_private_inventory_structure():
    """Force fix the private inventory structure to have only the correct columns"""
    private_path = os.path.join(os.path.dirname(__file__), '..', 'Private_Inventory.xlsx')
    
    try:
        if os.path.exists(private_path):
            # Read existing data
            df = pd.read_excel(private_path)
            
            # Define the correct columns
            required_columns = ['chemical_name', 'alias', 'cas_number', 'molecular_weight', 'smiles', 'barcode']
            
            # Create a new DataFrame with only the required columns
            new_df = pd.DataFrame()
            
            # Copy data from existing columns if they exist
            for col in required_columns:
                if col in df.columns:
                    new_df[col] = df[col]
                else:
                    new_df[col] = ''
            
            # Save the corrected structure
            new_df.to_excel(private_path, index=False)
            
            return jsonify({'message': 'Private inventory structure fixed successfully'}), 200
        else:
            return jsonify({'message': 'No private inventory file found'}), 404
    except Exception as e:
        return jsonify({'error': f'Failed to fix structure: {str(e)}'}), 500

@app.route('/api/inventory/private/check', methods=['POST'])
def check_private_inventory():
    """Check if a chemical exists in private inventory by name, alias, CAS, or SMILES"""
    chemical = request.json
    private_path = os.path.join(os.path.dirname(__file__), '..', 'Private_Inventory.xlsx')
    
    if not os.path.exists(private_path):
        return jsonify({'exists': False}), 200
    
    try:
        df = pd.read_excel(private_path)
        
        # Check for matches by name, alias, CAS, or SMILES
        name_match = df['chemical_name'].str.lower() == chemical.get('name', '').lower()
        alias_match = df['alias'].str.lower() == chemical.get('alias', '').lower()
        cas_match = df['cas_number'].astype(str) == str(chemical.get('cas', ''))
        smiles_match = df['smiles'].astype(str).str.lower() == str(chemical.get('smiles', '')).lower()
        
        exists = (name_match | alias_match | cas_match | smiles_match).any()
        return jsonify({'exists': bool(exists)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/experiment/context', methods=['GET', 'POST'])
def experiment_context():
    """Get or update experiment context"""
    global current_experiment
    
    if request.method == 'POST':
        current_experiment['context'] = request.json
        return jsonify({'message': 'Context updated'})
    
    return jsonify(current_experiment['context'])

@app.route('/api/experiment/materials', methods=['GET', 'POST'])
def experiment_materials():
    """Get or update experiment materials"""
    global current_experiment
    
    if request.method == 'POST':
        current_experiment['materials'] = request.json
        return jsonify({'message': 'Materials updated'})
    
    return jsonify(current_experiment['materials'])

@app.route('/api/experiment/procedure', methods=['GET', 'POST'])
def experiment_procedure():
    """Get or update experiment procedure (96-well plate)"""
    global current_experiment
    
    if request.method == 'POST':
        current_experiment['procedure'] = request.json
        return jsonify({'message': 'Procedure updated'})
    
    return jsonify(current_experiment['procedure'])

@app.route('/api/experiment/procedure-settings', methods=['GET', 'POST'])
def experiment_procedure_settings():
    """Get or update experiment procedure settings (reaction conditions and analytical details)"""
    global current_experiment
    
    if request.method == 'POST':
        current_experiment['procedure_settings'] = request.json
        return jsonify({'message': 'Procedure settings updated'})
    
    return jsonify(current_experiment.get('procedure_settings', {
        'reactionConditions': {
            'temperature': '',
            'time': '',
            'pressure': '',
            'wavelength': '',
            'remarks': ''
        },
        'analyticalDetails': {
            'uplcNumber': '',
            'method': '',
            'duration': '',
            'remarks': ''
        }
    }))

@app.route('/api/experiment/analytical', methods=['GET', 'POST'])
def experiment_analytical():
    """Get or update analytical data"""
    global current_experiment
    
    if request.method == 'POST':
        # Handle selected compounds update
        if 'selectedCompounds' in request.json:
            if 'analytical_data' not in current_experiment:
                current_experiment['analytical_data'] = {}
            current_experiment['analytical_data']['selectedCompounds'] = request.json['selectedCompounds']
            return jsonify({'message': 'Selected compounds updated'})
        else:
            # Handle other analytical data updates
            current_experiment['analytical_data'] = request.json
            return jsonify({'message': 'Analytical data updated'})
    
    # Return the analytical data structure that frontend expects
    analytical_data = current_experiment.get('analytical_data', {})
    if isinstance(analytical_data, list):
        # If it's a list (old format), convert to new format
        return jsonify({
            'selectedCompounds': [],
            'uploadedFiles': analytical_data
        })
    else:
        # Return the analytical data as is
        return jsonify(analytical_data)

@app.route('/api/experiment/analytical/template', methods=['POST'])
def export_analytical_template():
    try:
        data = request.get_json()
        compounds = data.get('compounds', [])
        
        if not compounds:
            return jsonify({'error': 'No compounds provided'}), 400
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytical Data"
        
        # Add headers - Well, Sample ID, then compound columns
        headers = ['Well', 'Sample ID']
        for i, compound in enumerate(compounds, 1):
            headers.extend([f'Name_{i}', f'Area_{i}'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Generate 96 wells (8 rows x 12 columns)
        rows = ["A", "B", "C", "D", "E", "F", "G", "H"]
        columns = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]
        
        # Get ELN number from context for sample IDs
        context = current_experiment.get('context', {})
        eln_number = context.get('eln_number', 'ELN-001')
        
        row_num = 2
        for row in rows:
            for col in columns:
                well = f"{row}{col}"
                sample_id = f"{eln_number}-{well}"
                
                # Add well and sample ID
                ws.cell(row=row_num, column=1, value=well)
                ws.cell(row=row_num, column=2, value=sample_id)
                
                # Add compound name and area columns
                for compound_idx, compound in enumerate(compounds):
                    # Cpd_(number)_Name column
                    ws.cell(row=row_num, column=3 + (compound_idx * 2), value=compound)
                    # Cmpd_(number)_area column (empty for user to fill)
                    ws.cell(row=row_num, column=4 + (compound_idx * 2), value="")
                
                row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Read the file and return it
        with open(tmp_file_path, 'rb') as f:
            file_content = f.read()
        
        # Clean up temporary file
        os.unlink(tmp_file_path)
        
        return send_file(
            io.BytesIO(file_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Analytical_Template_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/experiment/analytical/upload', methods=['POST'])
def upload_analytical_data():
    try:
        print("Upload endpoint called")
        
        if 'file' not in request.files:
            print("No file in request.files")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        print(f"File received: {file.filename}")
        
        if file.filename == '':
            print("Empty filename")
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        print(f"File extension: {file_ext}")
        
        if file_ext not in allowed_extensions:
            return jsonify({'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'}), 400
        
        # Read the file based on its type
        try:
            print("Attempting to read file")
            if file_ext == '.csv':
                df = pd.read_csv(file)
            else:  # Excel files
                df = pd.read_excel(file)
            print(f"File read successfully. Shape: {df.shape}")
        except Exception as e:
            print(f"Error reading file: {str(e)}")
            return jsonify({'error': f'Error reading file: {str(e)}'}), 400
        
        # Basic validation - check if file has expected structure
        if len(df.columns) < 2:
            print(f"File has insufficient columns: {len(df.columns)}")
            return jsonify({'error': 'File must have at least 2 columns (Well and Sample ID)'}), 400
        
        # Validate area columns (columns containing "Area_" in their names)
        area_columns = [col for col in df.columns if col.startswith('Area_')]
        print(f"Found area columns: {area_columns}")
        
        # Pre-process area columns: replace empty cells with 0
        for col in area_columns:
            # First, replace various empty representations with NaN
            df[col] = df[col].replace(['', ' ', 'nan', 'NaN', 'None', None], pd.NA)
            # Fill NaN values with 0
            df[col] = df[col].fillna(0)
            print(f"Pre-processed column {col}: replaced empty cells with 0")
        
        # Check if area columns contain only numerical data
        invalid_area_columns = []
        for col in area_columns:
            try:
                # Convert to numeric, coercing errors to NaN
                numeric_col = pd.to_numeric(df[col], errors='coerce')
                # Check if there are any NaN values (indicating non-numeric data)
                # After our pre-processing, NaN should only occur for truly non-numeric data
                if numeric_col.isna().any():
                    invalid_area_columns.append(col)
                    print(f"Column {col} contains non-numeric data")
                    # Show which values couldn't be converted
                    nan_mask = numeric_col.isna()
                    problematic_values = df.loc[nan_mask, col].unique()
                    print(f"Problematic values in {col}: {problematic_values}")
            except Exception as e:
                invalid_area_columns.append(col)
                print(f"Error validating column {col}: {str(e)}")
        
        if invalid_area_columns:
            return jsonify({
                'error': f'Area columns must contain only numerical data. Invalid columns: {", ".join(invalid_area_columns)}'
            }), 400
        
        print(f"Current experiment state before upload: {list(current_experiment.keys())}")
        
        # Store the uploaded data in the current experiment
        uploaded_data = {
            'filename': file.filename,
            'upload_date': datetime.now().isoformat(),
            'data': df.to_dict('records'),
            'columns': df.columns.tolist(),
            'shape': df.shape,
            'area_columns': area_columns  # Add area columns information
        }
        
        # Add to analytical data without overwriting existing data
        if 'analytical_data' not in current_experiment:
            current_experiment['analytical_data'] = {}
        
        # If analytical_data is a list (old format), convert it to new format
        if isinstance(current_experiment['analytical_data'], list):
            old_uploads = current_experiment['analytical_data']
            current_experiment['analytical_data'] = {
                'selectedCompounds': [],
                'uploadedFiles': old_uploads
            }
        
        if 'uploadedFiles' not in current_experiment['analytical_data']:
            current_experiment['analytical_data']['uploadedFiles'] = []
        
        current_experiment['analytical_data']['uploadedFiles'].append(uploaded_data)
        
        print(f"Upload successful. Current experiment keys: {list(current_experiment.keys())}")
        
        return jsonify({
            'message': 'File uploaded successfully',
            'filename': file.filename,
            'rows': len(df),
            'columns': len(df.columns),
            'area_columns': area_columns
        }), 200
        
    except Exception as e:
        print(f"Unexpected error in upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@app.route('/api/experiment/materials/upload', methods=['POST'])
def upload_materials_from_excel():
    """Upload materials from Excel file"""
    try:
        print("Materials upload endpoint called")
        
        # Ensure inventory is loaded
        if inventory_data is None:
            if not load_inventory():
                print("Warning: Could not load inventory data")
        
        if 'file' not in request.files:
            print("No file in request.files")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        print(f"File received: {file.filename}")
        
        if file.filename == '':
            print("Empty filename")
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = {'.xlsx', '.xls'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        print(f"File extension: {file_ext}")
        
        if file_ext not in allowed_extensions:
            return jsonify({'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'}), 400
        
        # Read the Excel file
        try:
            print("Attempting to read Excel file")
            excel_file = pd.ExcelFile(file)
            print(f"Excel sheets: {excel_file.sheet_names}")
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 400
        
        # Look for Materials sheet
        if 'Materials' not in excel_file.sheet_names:
            return jsonify({'error': 'No "Materials" sheet found in the Excel file'}), 400
        
        # Read the Materials sheet
        try:
            materials_df = pd.read_excel(file, sheet_name='Materials')
            print(f"Materials sheet read successfully. Shape: {materials_df.shape}")
        except Exception as e:
            print(f"Error reading Materials sheet: {str(e)}")
            return jsonify({'error': f'Error reading Materials sheet: {str(e)}'}), 400
        
        # Extract materials from the sheet
        materials = []
        for index, row in materials_df.iterrows():
            # Skip empty rows
            if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                continue
            
            # Extract material data based on expected columns (support both old and new column names)
            # Helper function to clean and validate field values
            def clean_field(value):
                if pd.isna(value):
                    return ''
                str_value = str(value).strip()
                # Return empty string for common "empty" representations
                if str_value.lower() in ['nan', 'null', 'none', '']:
                    return ''
                return str_value
            
            material = {
                'name': clean_field(row.get('chemical_name', row.get('Chemical_Name', row.get('Name', row.iloc[1] if len(row) > 1 else '')))),
                'alias': clean_field(row.get('alias', row.get('Alias', ''))),
                'cas': clean_field(row.get('cas_number', row.get('CAS_Number', row.get('CAS', '')))),
                'smiles': clean_field(row.get('smiles', row.get('SMILES', ''))),
                'molecular_weight': clean_field(row.get('molecular_weight', row.get('Molecular_Weight', row.get('Molecular Weight', '')))),
                'barcode': clean_field(row.get('barcode', row.get('Barcode', row.get('Lot number', '')))),
                'role': clean_field(row.get('role', row.get('Role', ''))),
                'source': 'excel_upload'
            }
            
            # Only add if name is not empty
            if material['name'] and material['name'] != 'nan':
                materials.append(material)
        
        if not materials:
            return jsonify({'error': 'No valid materials found in the Materials sheet'}), 400
        
        print(f"Extracted {len(materials)} materials from Excel file")
        for i, mat in enumerate(materials):
            print(f"  Material {i+1}: {mat['name']} - CAS: {mat['cas']} - Alias: {mat['alias']}")
        
        # Get current materials
        current_materials = current_experiment.get('materials', [])
        
        # Check for duplicates and add new materials
        added_materials = []
        skipped_materials = []
        
        # First, check all materials against the original current_materials list
        for material in materials:
            print(f"Processing material: {material['name']} (CAS: {material['cas']})")
            # Check if material already exists in current experiment (by name, CAS, or SMILES)
            is_duplicate = any(
                existing['name'] == material['name'] or
                (existing.get('cas') and material.get('cas') and existing['cas'] == material['cas']) or
                (existing.get('smiles') and material.get('smiles') and existing['smiles'] == material['smiles'])
                for existing in current_materials
            )
            
            if is_duplicate:
                print(f"  -> Skipping {material['name']} (duplicate)")
                skipped_materials.append(material['alias'] or material['name'])
            else:
                # Check if material exists in inventory or private inventory by CAS number
                inventory_material = None
                if material.get('cas') and material['cas'].strip():
                    # Check main inventory
                    if inventory_data is not None:
                        inventory_match = inventory_data[
                            inventory_data['cas_number'].astype(str).str.strip() == material['cas'].strip()
                        ]
                        if not inventory_match.empty:
                            inventory_material = inventory_match.iloc[0].to_dict()
                            print(f"  -> Found in main inventory: {inventory_material.get('chemical_name')}")
                        else:
                            print(f"  -> Not found in main inventory")
                    
                    # Check private inventory if not found in main inventory
                    if inventory_material is None:
                        private_path = os.path.join(os.path.dirname(__file__), '..', 'Private_Inventory.xlsx')
                        if os.path.exists(private_path):
                            try:
                                private_df = pd.read_excel(private_path, parse_dates=False)
                                # Convert all columns to string to avoid NaTType issues
                                for col in private_df.columns:
                                    private_df[col] = private_df[col].astype(str)
                                    private_df[col] = private_df[col].replace('nan', None)
                                
                                private_match = private_df[
                                    private_df['cas_number'].astype(str).str.strip() == material['cas'].strip()
                                ]
                                if not private_match.empty:
                                    inventory_material = private_match.iloc[0].to_dict()
                                    print(f"  -> Found in private inventory: {inventory_material.get('chemical_name')}")
                                else:
                                    print(f"  -> Not found in private inventory")
                            except Exception as e:
                                print(f"Error checking private inventory: {e}")
                
                # Use inventory data if found, otherwise use uploaded data
                if inventory_material:
                    print(f"  -> Using inventory data for {material['name']}")
                    # Use inventory data but preserve some uploaded data
                    final_material = {
                        'name': inventory_material.get('chemical_name', material['name']),
                        'alias': material.get('alias', inventory_material.get('alias', '')),
                        'cas': inventory_material.get('cas_number', material['cas']),
                        'smiles': inventory_material.get('smiles', material['smiles']),
                        'molecular_weight': inventory_material.get('molecular_weight', material['molecular_weight']),
                        'barcode': material.get('barcode', inventory_material.get('barcode', '')),
                        'role': material.get('role', ''),
                        'source': 'inventory_match',
                        'inventory_location': inventory_material.get('location', ''),
                        'supplier': inventory_material.get('supplier', '')
                    }
                else:
                    print(f"  -> Using uploaded data for {material['name']}")
                    # Use uploaded data
                    final_material = material.copy()
                    final_material['source'] = 'excel_upload'
                
                added_materials.append(final_material)
        
        # Then, add all new materials to the current_materials list at once
        current_materials.extend(added_materials)
        
        # Update the experiment materials
        current_experiment['materials'] = current_materials
        
        # Count materials by source
        inventory_matches = len([m for m in added_materials if m.get('source') == 'inventory_match'])
        excel_uploads = len([m for m in added_materials if m.get('source') == 'excel_upload'])
        
        print(f"Final results: Added {len(added_materials)} materials ({inventory_matches} from inventory, {excel_uploads} from upload data)")
        for i, mat in enumerate(added_materials):
            print(f"  Added {i+1}: {mat['name']} (source: {mat['source']})")
        
        return jsonify({
            'message': 'Materials uploaded successfully',
            'filename': file.filename,
            'total_materials': len(materials),
            'added_materials': len(added_materials),
            'skipped_materials': len(skipped_materials),
            'inventory_matches': inventory_matches,
            'excel_uploads': excel_uploads,
            'added_material_names': [m['alias'] or m['name'] for m in added_materials],
            'skipped_material_names': skipped_materials
        }), 200
        
    except Exception as e:
        print(f"Unexpected error in materials upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Materials upload failed: {str(e)}'}), 500

@app.route('/api/experiment/kit/analyze', methods=['POST'])
def analyze_kit():
    """Analyze kit Excel file and return materials and design data"""
    try:
        print("Kit analyze endpoint called")
        
        if 'file' not in request.files:
            print("No file in request.files")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        print(f"File received: {file.filename}")
        
        if file.filename == '':
            print("Empty filename")
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = {'.xlsx', '.xls'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        print(f"File extension: {file_ext}")
        
        if file_ext not in allowed_extensions:
            return jsonify({'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'}), 400
        
        # Read the Excel file
        try:
            print("Attempting to read Excel file")
            excel_file = pd.ExcelFile(file)
            print(f"Excel sheets: {excel_file.sheet_names}")
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 400
        
        # Look for Materials sheet
        if 'Materials' not in excel_file.sheet_names:
            return jsonify({'error': 'No "Materials" sheet found in the Excel file'}), 400
        
        # Look for Design sheet
        if 'Design' not in excel_file.sheet_names:
            return jsonify({'error': 'No "Design" sheet found in the Excel file'}), 400
        
        # Read the Materials sheet
        try:
            materials_df = pd.read_excel(file, sheet_name='Materials')
            print(f"Materials sheet read successfully. Shape: {materials_df.shape}")
        except Exception as e:
            print(f"Error reading Materials sheet: {str(e)}")
            return jsonify({'error': f'Error reading Materials sheet: {str(e)}'}), 400
        
        # Read the Design sheet
        try:
            design_df = pd.read_excel(file, sheet_name='Design')
            print(f"Design sheet read successfully. Shape: {design_df.shape}")
        except Exception as e:
            print(f"Error reading Design sheet: {str(e)}")
            return jsonify({'error': f'Error reading Design sheet: {str(e)}'}), 400
        
        # Extract materials from the Materials sheet
        materials = []
        for index, row in materials_df.iterrows():
            # Skip empty rows
            if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                continue
            
            # Extract material data based on expected columns (support both old and new column names)
            # Helper function to clean and validate field values
            def clean_field(value):
                if pd.isna(value):
                    return ''
                str_value = str(value).strip()
                # Return empty string for common "empty" representations
                if str_value.lower() in ['nan', 'null', 'none', '']:
                    return ''
                return str_value
            
            material = {
                'name': clean_field(row.get('chemical_name', row.get('Chemical_Name', row.get('Name', row.iloc[1] if len(row) > 1 else '')))),
                'alias': clean_field(row.get('alias', row.get('Alias', ''))),
                'cas': clean_field(row.get('cas_number', row.get('CAS_Number', row.get('CAS', '')))),
                'smiles': clean_field(row.get('smiles', row.get('SMILES', ''))),
                'molecular_weight': clean_field(row.get('molecular_weight', row.get('Molecular_Weight', row.get('Molecular Weight', '')))),
                'barcode': clean_field(row.get('barcode', row.get('Barcode', row.get('Lot number', '')))),
                'role': clean_field(row.get('role', row.get('Role', ''))),
                'source': 'kit_upload'
            }
            
            # Only add if name is not empty
            if material['name'] and material['name'] != 'nan':
                materials.append(material)
        
        if not materials:
            return jsonify({'error': 'No valid materials found in the Materials sheet'}), 400
        
        # Extract design data from the Design sheet
        design_data = {}
        kit_wells = set()
        
        for index, row in design_df.iterrows():
            # Skip empty rows
            if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                continue
            
            well = str(row.get('Well', row.iloc[0])).strip()
            if not well or well == 'nan':
                continue
            
            kit_wells.add(well)
            
            # Extract compounds and amounts from the row
            well_materials = []
            
            # Look for compound columns (e.g., "Compound 1 name", "Compound 1 amount")
            col_index = 2  # Start after Well and ID columns
            while col_index < len(row):
                if col_index + 1 < len(row):
                    compound_name = str(row.iloc[col_index]).strip()
                    compound_amount = str(row.iloc[col_index + 1]).strip()
                    
                    if compound_name and compound_name != 'nan' and compound_amount and compound_amount != 'nan':
                        # Find the material in our materials list
                        material = next((m for m in materials if m['name'] == compound_name or m['alias'] == compound_name), None)
                        if material:
                            well_materials.append({
                                'name': material['name'],
                                'alias': material['alias'],
                                'amount': compound_amount,
                                'unit': 'μmol'  # Default unit
                            })
                
                col_index += 2  # Move to next compound pair
            
            if well_materials:
                design_data[well] = well_materials
        
        # Determine kit size based on wells that actually have content
        if not design_data:
            return jsonify({'error': 'No wells with materials found in the Design sheet'}), 400
        
        # Use only wells that have materials for kit size calculation
        content_wells = list(design_data.keys())
        
        # Parse well positions to determine kit dimensions
        rows = set()
        cols = set()
        
        for well in content_wells:
            if len(well) >= 2:
                row_letter = well[0]
                col_number = int(well[1:])
                rows.add(row_letter)
                cols.add(col_number)
        
        # Convert row letters to numbers for calculation
        row_numbers = [ord(r) - ord('A') + 1 for r in rows]
        
        kit_size = {
            'rows': len(rows),
            'columns': len(cols),
            'total_wells': len(content_wells),
            'row_range': f"{min(rows)}-{max(rows)}" if len(rows) > 1 else min(rows),
            'col_range': f"{min(cols)}-{max(cols)}" if len(cols) > 1 else str(min(cols)),
            'wells': sorted(list(content_wells))
        }
        
        print(f"Kit analysis complete: {len(materials)} materials, {len(design_data)} wells with content")
        print(f"Kit size calculated: rows={len(rows)}, columns={len(cols)}, wells={sorted(content_wells)}")
        print(f"Row set: {rows}, Column set: {cols}")
        print(f"Content wells: {sorted(content_wells)}")
        
        return jsonify({
            'materials': materials,
            'design': design_data,
            'kit_size': kit_size,
            'filename': file.filename
        }), 200
        
    except Exception as e:
        print(f"Unexpected error in kit analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Kit analysis failed: {str(e)}'}), 500

@app.route('/api/experiment/kit/apply', methods=['POST'])
def apply_kit():
    """Apply kit to experiment with specified positioning"""
    global current_experiment
    
    try:
        print("Kit apply endpoint called")
        
        data = request.json
        materials = data.get('materials', [])
        design = data.get('design', {})
        position = data.get('position', '')
        kit_size = data.get('kit_size', {})
        destination_plate = data.get('destination_plate', '96')
        
        if not materials or not design or not position:
            return jsonify({'error': 'Missing required data: materials, design, or position'}), 400
        
        print(f"Applying kit with position: {position} on {destination_plate}-well plate")
        
        # Get current experiment data
        current_materials = current_experiment.get('materials', [])
        current_procedure = current_experiment.get('procedure', [])
        
        # Add materials to experiment (avoiding duplicates)
        added_materials = []
        skipped_materials = []
        
        for material in materials:
            # Check if material already exists (by name, CAS, or SMILES)
            is_duplicate = any(
                existing['name'] == material['name'] or
                (existing.get('cas') and material.get('cas') and existing['cas'] == material['cas']) or
                (existing.get('smiles') and material.get('smiles') and existing['smiles'] == material['smiles'])
                for existing in current_materials
            )
            
            if is_duplicate:
                skipped_materials.append(material['alias'] or material['name'])
            else:
                added_materials.append(material)
                current_materials.append(material)
        
        # Apply design to procedure based on position
        new_procedure_data = apply_kit_design_to_procedure(design, position, kit_size, current_procedure, destination_plate)
        
        # Update experiment
        current_experiment['materials'] = current_materials
        current_experiment['procedure'] = new_procedure_data
        
        return jsonify({
            'message': 'Kit applied successfully',
            'added_materials': len(added_materials),
            'skipped_materials': len(skipped_materials),
            'procedure_wells_updated': len([w for w in new_procedure_data if any(m['source'] == 'kit_upload' for m in w.get('materials', []))])
        }), 200
        
    except Exception as e:
        print(f"Unexpected error in kit application: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Kit application failed: {str(e)}'}), 500

def apply_kit_design_to_procedure(design, position, kit_size, current_procedure, destination_plate='96'):
    """Apply kit design to procedure data with position mapping"""
    
    # Create a dictionary for easier lookup of existing procedure data
    procedure_dict = {item['well']: item for item in current_procedure}
    
    # Calculate position mappings based on position parameter
    well_mappings = calculate_well_mappings(position, kit_size, destination_plate)
    
    # Apply design to mapped wells
    for kit_well, materials in design.items():
        if kit_well in well_mappings:
            target_wells = well_mappings[kit_well]
            
            for target_well in target_wells:
                # Get or create procedure entry for target well
                if target_well not in procedure_dict:
                    procedure_dict[target_well] = {
                        'well': target_well,
                        'materials': []
                    }
                
                # Add materials to the well
                for material in materials:
                    # Mark materials as coming from kit
                    kit_material = material.copy()
                    kit_material['source'] = 'kit_upload'
                    
                    # Check if this material already exists in this well
                    existing_material = next(
                        (m for m in procedure_dict[target_well]['materials'] 
                         if m['name'] == kit_material['name']),
                        None
                    )
                    
                    if existing_material:
                        # Add amounts if material already exists
                        try:
                            existing_amount = float(existing_material.get('amount', 0))
                            new_amount = float(kit_material.get('amount', 0))
                            existing_material['amount'] = str(existing_amount + new_amount)
                        except:
                            existing_material['amount'] = kit_material.get('amount', '0')
                    else:
                        # Add new material to well
                        procedure_dict[target_well]['materials'].append(kit_material)
    
    # Convert back to list format
    return list(procedure_dict.values())

def calculate_well_mappings(position, kit_size, destination_plate='96'):
    """Calculate well mappings based on position and kit size"""
    kit_wells = kit_size.get('wells', [])
    mappings = {}
    
    print(f"Calculating well mappings for position: {position}")
    print(f"Destination plate: {destination_plate}")
    print(f"Kit size: {kit_size}")
    print(f"Kit wells: {kit_wells}")
    
    if not kit_wells:
        return mappings
    
    # Handle new flexible positioning format
    if isinstance(position, dict):
        return calculate_flexible_well_mappings(position, kit_size, destination_plate)
    
    # Parse kit wells to understand the layout
    kit_rows = set()
    kit_cols = set()
    
    for well in kit_wells:
        if len(well) >= 2:
            row_letter = well[0]
            col_number = int(well[1:])
            kit_rows.add(row_letter)
            kit_cols.add(col_number)
    
    kit_rows = sorted(list(kit_rows))
    kit_cols = sorted(list(kit_cols))
    
    print(f"Parsed kit rows: {kit_rows}")
    print(f"Parsed kit cols: {kit_cols}")
    
    # Handle different position types
    if position == "top-left":
        # Map to A1-D6 region
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            # Calculate offset within kit
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            # Map to target position
            target_row = chr(ord('A') + row_offset)
            target_col = 1 + col_offset
            
            mappings[well] = [f"{target_row}{target_col}"]
    
    elif position == "top-right":
        # Map to A7-D12 region
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_row = chr(ord('A') + row_offset)
            target_col = 7 + col_offset
            
            mappings[well] = [f"{target_row}{target_col}"]
    
    elif position == "bottom-left":
        # Map to E1-H6 region
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_row = chr(ord('E') + row_offset)
            target_col = 1 + col_offset
            
            mappings[well] = [f"{target_row}{target_col}"]
    
    elif position == "bottom-right":
        # Map to E7-H12 region
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_row = chr(ord('E') + row_offset)
            target_col = 7 + col_offset
            
            mappings[well] = [f"{target_row}{target_col}"]
    
    elif position == "all-quadrants":
        # Map to all 4 quadrants
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_wells = []
            # Top-left
            target_wells.append(f"{chr(ord('A') + row_offset)}{1 + col_offset}")
            # Top-right
            target_wells.append(f"{chr(ord('A') + row_offset)}{7 + col_offset}")
            # Bottom-left
            target_wells.append(f"{chr(ord('E') + row_offset)}{1 + col_offset}")
            # Bottom-right
            target_wells.append(f"{chr(ord('E') + row_offset)}{7 + col_offset}")
            
            mappings[well] = target_wells
    
    elif position.startswith("row-"):
        # Full row positioning
        target_row = position.split("-")[1]
        print(f"Row positioning: target_row = {target_row}")
        for well in kit_wells:
            # Extract column number from well (e.g., "A1" -> 1, "A12" -> 12)
            col_number = int(well[1:])
            # Find the position of this column in the sorted kit columns
            col_offset = kit_cols.index(col_number)
            # Map to the target position starting from column 1
            target_col = 1 + col_offset
            print(f"Mapping well {well} (col {col_number}, offset {col_offset}) -> {target_row}{target_col}")
            mappings[well] = [f"{target_row}{target_col}"]
    
    elif position.startswith("col-"):
        # Full column positioning
        target_col = int(position.split("-")[1])
        for well in kit_wells:
            row_letter = well[0]
            row_offset = kit_rows.index(row_letter)
            target_row = chr(ord('A') + row_offset)
            mappings[well] = [f"{target_row}{target_col}"]
    
    elif position == "top-quadrants":
        # Map to top 2 quadrants (A1-D6 and A7-D12)
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_wells = []
            # Top-left
            target_wells.append(f"{chr(ord('A') + row_offset)}{1 + col_offset}")
            # Top-right
            target_wells.append(f"{chr(ord('A') + row_offset)}{7 + col_offset}")
            
            mappings[well] = target_wells
    
    elif position == "bottom-quadrants":
        # Map to bottom 2 quadrants (E1-H6 and E7-H12)
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_wells = []
            # Bottom-left
            target_wells.append(f"{chr(ord('E') + row_offset)}{1 + col_offset}")
            # Bottom-right
            target_wells.append(f"{chr(ord('E') + row_offset)}{7 + col_offset}")
            
            mappings[well] = target_wells
    
    elif position == "left-quadrants":
        # Map to left 2 quadrants (A1-D6 and E1-H6)
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_wells = []
            # Top-left
            target_wells.append(f"{chr(ord('A') + row_offset)}{1 + col_offset}")
            # Bottom-left
            target_wells.append(f"{chr(ord('E') + row_offset)}{1 + col_offset}")
            
            mappings[well] = target_wells
    
    elif position == "right-quadrants":
        # Map to right 2 quadrants (A7-D12 and E7-H12)
        for well in kit_wells:
            row_letter = well[0]
            col_number = int(well[1:])
            
            row_offset = kit_rows.index(row_letter)
            col_offset = kit_cols.index(col_number)
            
            target_wells = []
            # Top-right
            target_wells.append(f"{chr(ord('A') + row_offset)}{7 + col_offset}")
            # Bottom-right
            target_wells.append(f"{chr(ord('E') + row_offset)}{7 + col_offset}")
            
            mappings[well] = target_wells
    
    elif position.startswith("rows-"):
        # Multiple row positioning (e.g., "rows-A-D")
        parts = position.split("-")
        if len(parts) >= 3:
            start_row = parts[1]
            end_row = parts[2]
            start_row_idx = ord(start_row) - ord('A')
            
            for well in kit_wells:
                row_letter = well[0]
                col_number = int(well[1:])
                
                row_offset = kit_rows.index(row_letter)
                col_offset = kit_cols.index(col_number)
                
                target_row = chr(ord('A') + start_row_idx + row_offset)
                target_col = 1 + col_offset
                
                mappings[well] = [f"{target_row}{target_col}"]
    
    elif position.startswith("cols-"):
        # Multiple column positioning (e.g., "cols-1-6")
        parts = position.split("-")
        if len(parts) >= 3:
            start_col = int(parts[1])
            
            for well in kit_wells:
                row_letter = well[0]
                col_number = int(well[1:])
                
                row_offset = kit_rows.index(row_letter)
                col_offset = kit_cols.index(col_number)
                
                target_row = chr(ord('A') + row_offset)
                target_col = start_col + col_offset
                
                mappings[well] = [f"{target_row}{target_col}"]
    
    print(f"Final mappings: {mappings}")
    return mappings

def calculate_flexible_well_mappings(position_data, kit_size, destination_plate='96'):
    """Calculate well mappings for the new flexible positioning system"""
    kit_wells = kit_size.get('wells', [])
    mappings = {}
    
    if not kit_wells:
        return mappings
    
    strategy = position_data.get('strategy', 'exact_placement')
    positions = position_data.get('positions', [])
    kit_rows = position_data.get('kit_size', {}).get('rows', 1)
    kit_cols = position_data.get('kit_size', {}).get('cols', 1)
    
    print(f"Flexible positioning - Strategy: {strategy}")
    print(f"Positions: {positions}")
    print(f"Kit dimensions: {kit_rows}x{kit_cols}")
    
    # Get destination plate config
    plate_configs = {
        '24': {'rows': 4, 'cols': 6},
        '48': {'rows': 6, 'cols': 8}, 
        '96': {'rows': 8, 'cols': 12}
    }
    plate_config = plate_configs.get(destination_plate, plate_configs['96'])
    
    if strategy == 'exact_placement':
        # Kit matches plate exactly or default A1 placement
        start_row = 'A'
        start_col = 1
        
        for well in kit_wells:
            if len(well) >= 2:
                kit_row_letter = well[0]
                kit_col_number = int(well[1:])
                
                # Calculate offset from kit origin
                kit_row_offset = ord(kit_row_letter) - ord('A')
                kit_col_offset = kit_col_number - 1
                
                # Map to destination
                dest_row = chr(ord(start_row) + kit_row_offset)
                dest_col = start_col + kit_col_offset
                
                if dest_row <= chr(ord('A') + plate_config['rows'] - 1) and dest_col <= plate_config['cols']:
                    mappings[well] = [f"{dest_row}{dest_col}"]
    
    elif strategy == 'row_placement':
        # Map kit to specific rows
        for position_id in positions:
            if position_id.startswith('row-'):
                target_start_row = position_id.split('-')[1]
                
                for well in kit_wells:
                    if len(well) >= 2:
                        kit_row_letter = well[0]
                        kit_col_number = int(well[1:])
                        
                        # Calculate kit row offset from kit's starting row
                        kit_row_offset = ord(kit_row_letter) - ord('A')
                        
                        # Map to destination starting from target_start_row
                        dest_row = chr(ord(target_start_row) + kit_row_offset)
                        dest_col = kit_col_number
                        
                        if ord(dest_row) <= ord('A') + plate_config['rows'] - 1 and dest_col <= plate_config['cols']:
                            if well not in mappings:
                                mappings[well] = []
                            mappings[well].append(f"{dest_row}{dest_col}")
    
    elif strategy == 'col_placement':
        # Map kit to specific columns
        for position_id in positions:
            if position_id.startswith('col-'):
                target_start_col = int(position_id.split('-')[1])
                
                for well in kit_wells:
                    if len(well) >= 2:
                        kit_row_letter = well[0]
                        kit_col_number = int(well[1:])
                        
                        # Calculate kit column offset from kit's starting column
                        kit_col_offset = kit_col_number - 1
                        
                        # Map to destination starting from target_start_col
                        dest_row = kit_row_letter
                        dest_col = target_start_col + kit_col_offset
                        
                        if ord(dest_row) <= ord('A') + plate_config['rows'] - 1 and dest_col <= plate_config['cols']:
                            if well not in mappings:
                                mappings[well] = []
                            mappings[well].append(f"{dest_row}{dest_col}")
    
    elif strategy == 'quadrant_placement':
        # Map kit to specific quadrants (adjust based on destination plate)
        if destination_plate == '48':
            # 48-well plate (6x8) - 4x6 kit can fit in two positions
            quadrant_offsets = {
                'top-left': (0, 0),      # A1 starting position (A1-D6)
                'top-right': (2, 0),     # C1 starting position (C1-F6) - shifted down 2 rows
            }
        else:  # 96-well plate
            quadrant_offsets = {
                'top-left': (0, 0),      # A1 starting position
                'top-right': (0, 6),     # A7 starting position  
                'bottom-left': (4, 0),   # E1 starting position
                'bottom-right': (4, 6)   # E7 starting position
            }
        
        for position_id in positions:
            if position_id in quadrant_offsets:
                row_offset, col_offset = quadrant_offsets[position_id]
                
                for well in kit_wells:
                    if len(well) >= 2:
                        kit_row_letter = well[0]
                        kit_col_number = int(well[1:])
                        
                        # Calculate offset from kit origin
                        kit_row_offset = ord(kit_row_letter) - ord('A')
                        kit_col_offset = kit_col_number - 1
                        
                        # Map to destination quadrant
                        dest_row = chr(ord('A') + row_offset + kit_row_offset)
                        dest_col = 1 + col_offset + kit_col_offset
                        
                        if ord(dest_row) <= ord('A') + plate_config['rows'] - 1 and dest_col <= plate_config['cols']:
                            if well not in mappings:
                                mappings[well] = []
                            mappings[well].append(f"{dest_row}{dest_col}")
    
    elif strategy == 'row_pair_placement':
        # Map kit to specific row pairs (for 2x12 kits)
        row_pair_offsets = {
            'AB': 0,  # Start at row A
            'CD': 2,  # Start at row C  
            'EF': 4,  # Start at row E
            'GH': 6   # Start at row G
        }
        
        for position_id in positions:
            if position_id in row_pair_offsets:
                row_offset = row_pair_offsets[position_id]
                
                for well in kit_wells:
                    if len(well) >= 2:
                        kit_row_letter = well[0]
                        kit_col_number = int(well[1:])
                        
                        # Calculate offset from kit origin
                        kit_row_offset = ord(kit_row_letter) - ord('A')
                        
                        # Map to destination row pair
                        dest_row = chr(ord('A') + row_offset + kit_row_offset)
                        dest_col = kit_col_number
                        
                        if ord(dest_row) <= ord('A') + plate_config['rows'] - 1 and dest_col <= plate_config['cols']:
                            if well not in mappings:
                                mappings[well] = []
                            mappings[well].append(f"{dest_row}{dest_col}")
    
    elif strategy == 'half_placement':
        # Map kit to upper or lower half (for 4x12 kits)
        half_offsets = {
            'upper': 0,  # Start at row A
            'lower': 4   # Start at row E
        }
        
        for position_id in positions:
            if position_id in half_offsets:
                row_offset = half_offsets[position_id]
                
                for well in kit_wells:
                    if len(well) >= 2:
                        kit_row_letter = well[0]
                        kit_col_number = int(well[1:])
                        
                        # Calculate offset from kit origin
                        kit_row_offset = ord(kit_row_letter) - ord('A')
                        
                        # Map to destination half
                        dest_row = chr(ord('A') + row_offset + kit_row_offset)
                        dest_col = kit_col_number
                        
                        if ord(dest_row) <= ord('A') + plate_config['rows'] - 1 and dest_col <= plate_config['cols']:
                            if well not in mappings:
                                mappings[well] = []
                            mappings[well].append(f"{dest_row}{dest_col}")

    elif strategy == 'block_placement':
        # Map kit to specific blocks/quadrants
        blocks = position_data.get('blocks', [])
        
        for block in blocks:
            start_row = block.get('startRow', 'A')
            start_col = block.get('startCol', 1)
            
            for well in kit_wells:
                if len(well) >= 2:
                    kit_row_letter = well[0]
                    kit_col_number = int(well[1:])
                    
                    # Calculate offset from kit origin
                    kit_row_offset = ord(kit_row_letter) - ord('A')
                    kit_col_offset = kit_col_number - 1
                    
                    # Map to destination block
                    dest_row = chr(ord(start_row) + kit_row_offset)
                    dest_col = start_col + kit_col_offset
                    
                    if ord(dest_row) <= ord('A') + plate_config['rows'] - 1 and dest_col <= plate_config['cols']:
                        if well not in mappings:
                            mappings[well] = []
                        mappings[well].append(f"{dest_row}{dest_col}")
    
    print(f"Flexible mappings result: {mappings}")
    return mappings

@app.route('/api/experiment/procedure/update-plate-type', methods=['POST'])
def update_procedure_plate_type():
    """Update the procedure plate type when kit is applied"""
    global current_experiment
    
    try:
        data = request.json
        new_plate_type = data.get('plate_type', '96')
        current_procedure = data.get('current_procedure', [])
        
        print(f"Updating procedure plate type to: {new_plate_type}")
        
        # Store the plate type information in the experiment context
        if 'context' not in current_experiment:
            current_experiment['context'] = {}
        
        current_experiment['context']['plate_type'] = new_plate_type
        
        # Validate that all wells in current procedure fit in the new plate type
        plate_configs = {
            '24': {'max_row': 'D', 'max_col': 6},
            '48': {'max_row': 'F', 'max_col': 8},
            '96': {'max_row': 'H', 'max_col': 12}
        }
        
        plate_config = plate_configs.get(new_plate_type, plate_configs['96'])
        
        # Filter out any procedure entries that don't fit in the new plate
        valid_procedure = []
        for entry in current_procedure:
            well = entry.get('well', '')
            if well and len(well) >= 2:
                row_letter = well[0]
                col_number = int(well[1:]) if well[1:].isdigit() else 0
                
                if (row_letter <= plate_config['max_row'] and 
                    col_number <= plate_config['max_col']):
                    valid_procedure.append(entry)
        
        # Update the procedure with valid entries
        current_experiment['procedure'] = valid_procedure
        
        return jsonify({
            'success': True, 
            'plate_type': new_plate_type,
            'filtered_wells': len(current_procedure) - len(valid_procedure)
        })
        
    except Exception as e:
        print(f"Error updating procedure plate type: {str(e)}")
        return jsonify({'error': f'Failed to update plate type: {str(e)}'}), 500

@app.route('/api/experiment/results', methods=['GET', 'POST'])
def experiment_results():
    """Get or update experiment results"""
    global current_experiment
    
    if request.method == 'POST':
        current_experiment['results'] = request.json
        return jsonify({'message': 'Results updated'})
    
    return jsonify(current_experiment['results'])

@app.route('/api/experiment/export', methods=['POST'])
def export_experiment():
    """Export experiment data to Excel format"""
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Context sheet
    ws_context = wb.create_sheet("Context")
    context_data = [
        ['Author', current_experiment['context'].get('author', '')],
        ['Date', current_experiment['context'].get('date', '')],
        ['Project', current_experiment['context'].get('project', '')],
        ['ELN', current_experiment['context'].get('eln', '')],
        ['Objective', current_experiment['context'].get('objective', '')]
    ]
    
    for row in context_data:
        ws_context.append(row)
    
    # Materials sheet
    ws_materials = wb.create_sheet("Materials")
    if current_experiment.get('materials'):
        # Add headers - match inventory column names exactly (lowercase) and order
        headers = ['Nr', 'chemical_name', 'alias', 'cas_number', 'molecular_weight', 'smiles', 'barcode', 'role', 'source', 'supplier']
        ws_materials.append(headers)
        
        # Load inventory data to enrich materials
        inventory_enrichment = {}
        if inventory_data is not None:
            for _, inv_item in inventory_data.iterrows():
                # Create lookup keys for matching
                name_key = str(inv_item.get('chemical_name', '')).lower()
                cas_key = str(inv_item.get('cas_number', '')).lower()
                alias_key = str(inv_item.get('alias', '')).lower()
                
                # Store inventory data for matching
                inventory_enrichment[name_key] = inv_item.to_dict()
                if cas_key and cas_key != 'nan':
                    inventory_enrichment[cas_key] = inv_item.to_dict()
                if alias_key and alias_key != 'nan':
                    inventory_enrichment[alias_key] = inv_item.to_dict()
        
        # Add materials with enriched data from inventory
        for i, material in enumerate(current_experiment['materials'], 1):
            # Try to find matching inventory data
            enriched_data = {}
            material_name = str(material.get('name', '')).lower()
            material_cas = str(material.get('cas', '')).lower()
            material_alias = str(material.get('alias', '')).lower()
            
            # Look for matches in inventory
            if material_name in inventory_enrichment:
                enriched_data = inventory_enrichment[material_name]
            elif material_cas in inventory_enrichment and material_cas != 'nan':
                enriched_data = inventory_enrichment[material_cas]
            elif material_alias in inventory_enrichment and material_alias != 'nan':
                enriched_data = inventory_enrichment[material_alias]
            
            # Use material data first, then enrich with inventory data
            row = [
                i,
                material.get('name', ''),
                material.get('alias', enriched_data.get('alias', '')),
                material.get('cas', enriched_data.get('cas_number', '')),
                material.get('molecular_weight', enriched_data.get('molecular_weight', '')),
                material.get('smiles', enriched_data.get('smiles', '')),
                material.get('barcode', enriched_data.get('barcode', '')),
                material.get('role', ''),
                material.get('source', enriched_data.get('source', '')),
                material.get('supplier', enriched_data.get('supplier', ''))
            ]
            ws_materials.append(row)
    
    # Procedure sheet
    ws_procedure = wb.create_sheet("Procedure")
    if current_experiment.get('procedure'):
        # Add headers for 96-well plate
        headers = ['Nr', 'Well', 'ID']
        # Add compound columns (up to 15 compounds)
        for i in range(1, 16):
            headers.extend([f'Compound-{i}_name', f'Compound-{i}_mmol'])
        # Add reagent columns (up to 5 reagents)
        for i in range(1, 6):
            headers.extend([f'Reagent-{i}_name', f'Reagent-{i}_mmol'])
        # Add solvent columns (up to 3 solvents)
        for i in range(1, 4):
            headers.extend([f'Solvent-{i}_name', f'Solvent-{i}_uL'])
        
        ws_procedure.append(headers)
        
        # Add procedure data
        for i, well_data in enumerate(current_experiment['procedure'], 1):
            row = [i, well_data.get('well', ''), well_data.get('id', '')]
            
            # Add compounds
            for j in range(1, 16):
                row.extend([
                    well_data.get(f'compound_{j}_name', ''),
                    well_data.get(f'compound_{j}_mmol', '')
                ])
            
            # Add reagents
            for j in range(1, 6):
                row.extend([
                    well_data.get(f'reagent_{j}_name', ''),
                    well_data.get(f'reagent_{j}_mmol', '')
                ])
            
            # Add solvents
            for j in range(1, 4):
                row.extend([
                    well_data.get(f'solvent_{j}_name', ''),
                    well_data.get(f'solvent_{j}_ul', '')
                ])
            
            ws_procedure.append(row)
    
    # Analytical data sheet
    ws_analytical = wb.create_sheet("Analytical data (1)")
    if current_experiment.get('analytical_data'):
        # Add headers
        headers = ['Nr', 'Well', 'ID']
        for i in range(1, 16):
            headers.extend([f'Compound-{i}_name', f'Compound-{i}_area'])
        
        ws_analytical.append(headers)
        
        # Add analytical data
        for i, analytical_data in enumerate(current_experiment['analytical_data'], 1):
            row = [i, analytical_data.get('well', ''), analytical_data.get('id', '')]
            
            for j in range(1, 16):
                row.extend([
                    analytical_data.get(f'compound_{j}_name', ''),
                    analytical_data.get(f'compound_{j}_area', '')
                ])
            
            ws_analytical.append(row)
    
    # Results sheet
    ws_results = wb.create_sheet("Results (1)")
    if current_experiment.get('results'):
        # Add headers
        headers = ['Nr', 'Well', 'ID', 'Conversion_%', 'Yield_%', 'Selectivity_%']
        ws_results.append(headers)
        
        # Add results data
        for i, result_data in enumerate(current_experiment['results'], 1):
            row = [
                i,
                result_data.get('well', ''),
                result_data.get('id', ''),
                result_data.get('conversion_percent', ''),
                result_data.get('yield_percent', ''),
                result_data.get('selectivity_percent', '')
            ]
            ws_results.append(row)
    
    # Well Contents sheet - Detailed view of each well
    ws_well_contents = wb.create_sheet("Well Contents")
    
    # Create a mapping of materials by name for quick lookup
    materials_map = {}
    for material in current_experiment.get('materials', []):
        materials_map[material.get('name', '').lower()] = material
    
    # Initialize well contents data
    well_contents = {}
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for row in range(1, 13):
            well = f'{col}{row}'
            well_contents[well] = {
                'compounds': [],
                'reagents': [],
                'solvents': []
            }
    
    # Fill in well contents from procedure data
    if current_experiment.get('procedure'):
        for well_data in current_experiment['procedure']:
            well = well_data.get('well', '')
            if well in well_contents:
                # Process materials array
                materials = well_data.get('materials', [])
                
                for material in materials:
                    name = material.get('name', '')
                    amount = material.get('amount', '')
                    alias = material.get('alias', '')
                    cas = material.get('cas', '')
                    
                    if name and amount:
                        # For now, treat all materials as compounds
                        # You can add logic here to distinguish compounds, reagents, solvents
                        well_contents[well]['compounds'].append({
                            'name': name,
                            'amount': amount,
                            'alias': alias,
                            'cas': cas
                        })
    
    # Find the maximum number of compounds across all wells to determine column count
    max_compounds = 0
    for well in well_contents:
        compounds_count = len(well_contents[well]['compounds'])
        reagents_count = len(well_contents[well]['reagents'])
        solvents_count = len(well_contents[well]['solvents'])
        total_materials = compounds_count + reagents_count + solvents_count
        max_compounds = max(max_compounds, total_materials)
    
    # Create header row
    headers = ['Well']
    for i in range(1, max_compounds + 1):
        headers.extend([f'Compound_{i}_Name', f'Compound_{i}_Alias', f'Compound_{i}_CAS', f'Compound_{i}_Amount'])
    
    ws_well_contents.append(headers)
    
    # Add data for each well (all 96 wells)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for row in range(1, 13):
            well = f'{col}{row}'
            contents = well_contents[well]
            
            # Combine all materials into a single list
            all_materials = []
            all_materials.extend(contents['compounds'])
            all_materials.extend(contents['reagents'])
            all_materials.extend(contents['solvents'])
            
            # Create row data
            row_data = [well]
            
            # Add materials to columns (4 columns per material)
            for i in range(max_compounds):
                if i < len(all_materials):
                    material = all_materials[i]
                    row_data.extend([
                        material['name'],
                        material.get('alias', ''),
                        material.get('cas', ''),
                        material['amount']
                    ])
                else:
                    # Fill empty columns
                    row_data.extend(['', '', '', ''])
            
            ws_well_contents.append(row_data)
    
    # Procedure Settings sheet
    ws_procedure_settings = wb.create_sheet("Procedure")
    
    # Reaction Conditions section
    ws_procedure_settings.append(['Reaction Conditions'])
    ws_procedure_settings.append(['Parameter', 'Value', 'Unit'])
    ws_procedure_settings.append(['Temperature', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('temperature', ''), 'degC'])
    ws_procedure_settings.append(['Time', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('time', ''), 'h'])
    ws_procedure_settings.append(['Pressure', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('pressure', ''), 'bar'])
    ws_procedure_settings.append(['Wavelength', current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('wavelength', ''), 'nm'])
    ws_procedure_settings.append([''])  # Empty row for spacing
    ws_procedure_settings.append(['Remarks'])
    ws_procedure_settings.append([current_experiment.get('procedure_settings', {}).get('reactionConditions', {}).get('remarks', '')])
    
    # Analytical Details section
    ws_procedure_settings.append([''])  # Empty row for spacing
    ws_procedure_settings.append(['Analytical Details'])
    ws_procedure_settings.append(['Parameter', 'Value', 'Unit'])
    ws_procedure_settings.append(['UPLC #', current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('uplcNumber', ''), ''])
    ws_procedure_settings.append(['Method', current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('method', ''), ''])
    ws_procedure_settings.append(['Duration', current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('duration', ''), 'min'])
    ws_procedure_settings.append([''])  # Empty row for spacing
    ws_procedure_settings.append(['Remarks'])
    ws_procedure_settings.append([current_experiment.get('procedure_settings', {}).get('analyticalDetails', {}).get('remarks', '')])
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    return send_file(tmp_path, as_attachment=True, 
                    download_name=f'HTE_experiment_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@app.route('/api/molecule/image', methods=['POST'])
def get_molecule_image():
    """Generate molecule image from SMILES string"""
    data = request.get_json()
    smiles = data.get('smiles', '').strip()
    
    if not smiles:
        return jsonify({'error': 'SMILES string is required'}), 400
    
    # Get optional image size
    width = data.get('width', 300)
    height = data.get('height', 300)
    
    # Generate image
    image_data = generate_molecule_image(smiles, (width, height))
    
    if image_data is None:
        return jsonify({'error': 'Invalid SMILES string'}), 400
    
    return jsonify({
        'image': image_data,
        'format': 'png',
        'size': {'width': width, 'height': height}
    })

@app.route('/api/upload/sdf', methods=['POST'])
def upload_sdf():
    """Upload and parse SDF file"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith('.sdf'):
        return jsonify({'error': 'File must be in SDF format'}), 400
    
    try:
        # Read file content
        sdf_content = file.read().decode('utf-8')
        
        # Parse SDF file
        molecules = parse_sdf_file(sdf_content)
        
        if not molecules:
            return jsonify({'error': 'No valid molecules found in SDF file'}), 400
        
        # Assign ID-based names to all molecules
        for i, molecule in enumerate(molecules):
            molecule['name'] = f"ID-{(i+1):02d}"
        
        return jsonify({
            'molecules': molecules,
            'total_molecules': len(molecules)
        })
        
    except Exception as e:
        return jsonify({'error': f'Error processing SDF file: {str(e)}'}), 500

@app.route('/api/experiment/heatmap', methods=['GET', 'POST'])
def experiment_heatmap():
    """Handle heatmap data persistence"""
    global current_experiment
    
    if request.method == 'GET':
        return jsonify(current_experiment.get('heatmap_data', {}))
    
    elif request.method == 'POST':
        data = request.get_json()
        current_experiment['heatmap_data'] = data
        return jsonify({'message': 'Heatmap data saved successfully'})

@app.route('/api/experiment/reset', methods=['POST'])
def reset_experiment():
    """Reset current experiment"""
    global current_experiment
    current_experiment = {
        'context': {},
        'materials': [],
        'procedure': [],
        'analytical_data': [],
        'results': [],
        'heatmap_data': {}
    }
    return jsonify({'message': 'Experiment reset'})

if __name__ == '__main__':
    # Load inventory on startup
    load_inventory()
    app.run(debug=True, port=5000) 